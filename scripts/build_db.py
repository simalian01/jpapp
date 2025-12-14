#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
从你提供的 Excel（重点：红宝书 sheet）生成离线词库 sqlite。

输出 schema（核心）：
- items(id INTEGER PRIMARY KEY, deck TEXT, level TEXT, term TEXT, reading TEXT, sheet TEXT, search_text TEXT, data_json TEXT)
- media(id INTEGER PRIMARY KEY AUTOINCREMENT, item_id INTEGER, type TEXT, path TEXT)
- meta(key TEXT PRIMARY KEY, value TEXT)

说明：
- term：优先用「汉字/外文」，否则用「假名」
- reading：用「假名」
- level：优先从音频文件名里提取 n1~n5（如 “01 n5-あ (2).mp3” -> N5）
- image：来自「图源」列
- audio：来自「音源路径」列
"""
import argparse
import json
import re
import sqlite3
from pathlib import Path

import openpyxl

LEVEL_RE = re.compile(r'(?i)\bn([1-5])\b')

def norm(v):
    if v is None:
        return ""
    s = str(v).strip()
    return s

def extract_level(audio_path: str) -> str:
    m = LEVEL_RE.search(audio_path.replace("\\", "/"))
    if not m:
        return ""
    return f"N{m.group(1)}"

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", required=True, help="Excel 文件路径")
    ap.add_argument("--out", required=True, help="输出 sqlite 路径")
    args = ap.parse_args()

    xls = Path(args.excel)
    out = Path(args.out)
    if not xls.exists():
        raise SystemExit(f"Excel not found: {xls}")

    wb = openpyxl.load_workbook(xls, data_only=True)
    if "红宝书" not in wb.sheetnames:
        raise SystemExit("Excel 中找不到 sheet：红宝书")

    ws = wb["红宝书"]

    # header row 1
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    header_map = {}
    for i, h in enumerate(headers, start=1):
        if isinstance(h, str) and h.strip():
            header_map[h.strip()] = i

    # 需要的列
    col_order = header_map.get("顺序", 1)
    col_kana = header_map.get("假名")
    col_kanji = header_map.get("汉字/外文")
    col_img = header_map.get("图源")  # 图源出现两次，openpyxl 会取最后一次；我们后面再兼容
    # 兼容：图源可能有两个同名列
    img_cols = [i for i, h in enumerate(headers, start=1) if (isinstance(h, str) and h.strip() == "图源")]
    if img_cols:
        col_img = img_cols[-1]

    col_audio = header_map.get("音源路径")
    if col_audio is None:
        # 有些版本可能只叫“音源”
        audio_cols = [i for i, h in enumerate(headers, start=1) if (isinstance(h, str) and "音源" in h)]
        col_audio = audio_cols[-1] if audio_cols else None

    if col_kana is None or col_kanji is None or col_audio is None or col_img is None:
        raise SystemExit(f"列不完整：假名={col_kana}, 汉字/外文={col_kanji}, 图源={col_img}, 音源路径={col_audio}")

    # 建库
    if out.exists():
        out.unlink()

    conn = sqlite3.connect(out)
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA synchronous=NORMAL;")
    conn.execute("PRAGMA temp_store=MEMORY;")

    conn.executescript("""
    CREATE TABLE IF NOT EXISTS meta(
      key TEXT PRIMARY KEY,
      value TEXT
    );

    CREATE TABLE IF NOT EXISTS items(
      id INTEGER PRIMARY KEY,
      deck TEXT NOT NULL,
      level TEXT,
      term TEXT NOT NULL,
      reading TEXT,
      sheet TEXT NOT NULL,
      search_text TEXT,
      data_json TEXT
    );

    CREATE TABLE IF NOT EXISTS media(
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      item_id INTEGER NOT NULL,
      type TEXT NOT NULL,
      path TEXT NOT NULL
    );

    CREATE INDEX IF NOT EXISTS idx_items_deck_level ON items(deck, level);
    CREATE INDEX IF NOT EXISTS idx_items_term ON items(term);
    CREATE INDEX IF NOT EXISTS idx_media_item ON media(item_id);
    """)

    conn.execute("INSERT OR REPLACE INTO meta(key,value) VALUES(?,?)", ("source_excel", str(xls)))
    conn.execute("INSERT OR REPLACE INTO meta(key,value) VALUES(?,?)", ("generated_at", __import__("datetime").datetime.now().isoformat()))

    # 数据起始：红宝书在你的文件里通常从第3行开始（第2行有杂项）
    # 这里用“顺序”列是整数作为判断。
    rows = []
    for r in range(2, ws.max_row + 1):
        order_val = ws.cell(row=r, column=col_order).value
        if order_val is None:
            continue
        try:
            item_id = int(order_val)
        except Exception:
            continue

        kana = norm(ws.cell(row=r, column=col_kana).value)
        kanji = norm(ws.cell(row=r, column=col_kanji).value)
        term = kanji if kanji else kana
        if not term:
            continue

        img_path = norm(ws.cell(row=r, column=col_img).value)
        audio_path = norm(ws.cell(row=r, column=col_audio).value)

        level = extract_level(audio_path)
        deck = "红宝书"
        sheet = "红宝书"

        search_text = " ".join([term, kana, kanji]).strip()
        data = {
            "kana": kana,
            "kanji": kanji,
            "excel_row": r,
        }
        rows.append((item_id, deck, level, term, kana, sheet, search_text, json.dumps(data, ensure_ascii=False), img_path, audio_path))

    # 写入
    conn.execute("BEGIN;")
    conn.executemany(
        "INSERT OR REPLACE INTO items(id,deck,level,term,reading,sheet,search_text,data_json) VALUES(?,?,?,?,?,?,?,?)",
        [(rid, deck, level, term, reading, sheet, st, dj) for (rid, deck, level, term, reading, sheet, st, dj, _, _) in rows]
    )
    # media
    media_rows = []
    for (rid, deck, level, term, reading, sheet, st, dj, img_path, audio_path) in rows:
        if img_path:
            media_rows.append((rid, "image", img_path))
        if audio_path:
            media_rows.append((rid, "audio", audio_path))
    conn.executemany("INSERT INTO media(item_id,type,path) VALUES(?,?,?)", media_rows)
    conn.execute("COMMIT;")

    print(f"OK: {out} items={len(rows)} media={len(media_rows)}")
    conn.close()

if __name__ == "__main__":
    main()
